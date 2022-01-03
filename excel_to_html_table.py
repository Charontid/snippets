import openpyxl as pyxl
import sys


def create_testfile():
    wb = pyxl.Workbook()
    ws = wb.active
    ws.title = "Table1"
    for col in range(1, 5):
        ws.cell(row=1, column=col, value=f"H{col}")
    for row in range(1, 5):
        for col in range(1, 5):
            ws.cell(row=row+1, column=col, value=row+col)
    wb.save('test.xlsx')


def parse_row(ws, row, header=False):
    tmp = list()
    tmp.append('  <tr>')
    for col in range(1, ws.max_column+1):
        cur = list()
        cur.append('    <th>') if header else cur.append('    <td>')
        cur.append(str(ws.cell(row=row, column=col).value))
        cur.append('</th>') if header else cur.append('</td>')
        tmp.append(''.join(cur))
    tmp.append('  </tr>')
    return '\n'.join(tmp)


def parse_to_html_table(file, sheet, header):
    wb = pyxl.load_workbook(file)
    ws = wb[sheet]

    with open('out.html', 'w') as html_file:
        html_file.write("<table>\n")
        html_file.write(parse_row(ws, 1, header=header))
        html_file.write('\n')

        for row in range(2, ws.max_row+1):
            html_file.write(parse_row(ws, row))
            html_file.write('\n')
        html_file.write("</table>\n")


def main(file=None, sheet=None, header=False):
    if file is None:
        create_testfile()
        file = 'test.xlsx'
        sheet = 'Table1'
        header = True
    parse_to_html_table(file, sheet, header)


if __name__ == '__main__':
    """
    Missing input sanity-checks!
    """
    if len(sys.argv) == 3:
        main(file=sys.argv[1], sheet=sys.argv[2])
    elif len(sys.argv) == 4:
        main(file=sys.argv[1], sheet=sys.argv[2], header=sys.argv[3])
    else:
        main()
